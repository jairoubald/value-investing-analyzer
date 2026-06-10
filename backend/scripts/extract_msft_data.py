"""Extract 1 DATA sheet from FINANCIAL TOOL.xlsx into preload JSON."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

EXCEL_PATH = Path(
    r"c:\Users\HOME\Desktop\4 - FUND JAIRO\2. ASSETS\2.1 THESIS\0. THESIS TOOL\FINANCIAL TOOL.xlsx"
)
OUT_PATH = Path(__file__).resolve().parents[1] / "data" / "msft_1data.json"

MAX_ROW = 250
MAX_COL = 47


def cell_value(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return str(v)


def main() -> None:
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["1 DATA"]

    years = []
    for c in range(5, 17):
        years.append(
            {
                "col": c,
                "fy": cell_value(ws.cell(12, c).value),
                "end_date": cell_value(ws.cell(13, c).value),
            }
        )

    grid: dict[str, dict[str, object]] = {}
    for r in range(1, MAX_ROW + 1):
        row_obj: dict[str, object] = {}
        label = ws.cell(r, 3).value or ws.cell(r, 4).value
        if label:
            row_obj["label"] = str(label)
        values = {}
        for c in range(5, MAX_COL + 1):
            v = ws.cell(r, c).value
            if v is not None:
                values[str(c)] = cell_value(v)
        if values:
            row_obj["values"] = values
        if row_obj:
            grid[str(r)] = row_obj

    payload = {
        "ticker": "MSFT",
        "company": cell_value(ws["C5"].value),
        "currency": cell_value(ws["F5"].value),
        "units": cell_value(ws["H5"].value),
        "years": years,
        "grid": grid,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload), encoding="utf-8")
    print(f"Wrote {OUT_PATH} ({OUT_PATH.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
