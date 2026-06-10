"""Inspect 3 GRAPHS sheet: sections, chart titles, data ranges."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH = Path(
    r"c:\Users\HOME\Desktop\4 - FUND JAIRO\2. ASSETS\2.1 THESIS\0. THESIS TOOL\FINANCIAL TOOL.xlsx"
)
OUT = Path(__file__).resolve().parents[1] / "data" / "graphs_inventory.txt"


def main() -> None:
    wb = load_workbook(PATH, data_only=True)
    ws = wb["3 GRAPHS"]
    lines: list[str] = []

    lines.append(f"Sheet size: {ws.max_row} x {ws.max_column}")
    lines.append("")

    lines.append("=== SECTION HEADERS (column B) ===")
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if not v or not isinstance(v, str):
            continue
        s = v.strip()
        if s and s[0].isdigit() and "." in s[:4] or (s[0].isdigit() and " " in s[:6]):
            lines.append(f"Row {r}: {s}")

    lines.append("")
    lines.append("=== ROW LABELS NEAR CHART DATA (column C, sample) ===")
    for r in range(1, min(ws.max_row, 310) + 1):
        v = ws.cell(r, 3).value
        if v and isinstance(v, str) and len(v.strip()) > 2:
            if r <= 20 or (r >= 25 and r <= 50) or (r >= 130 and r <= 200) or (r >= 250):
                lines.append(f"C{r}: {v.strip()}")

    lines.append("")
    lines.append("=== EMBEDDED CHARTS ===")
    charts = list(getattr(ws, "_charts", []) or [])
    lines.append(f"Count: {len(charts)}")
    for i, ch in enumerate(charts, 1):
        title = "?"
        try:
            t = ch.title
            if t and hasattr(t, "text") and t.text:
                title = str(t.text)
            elif t and hasattr(t, "tx") and t.tx and hasattr(t.tx, "v") and t.tx.v:
                title = str(t.tx.v)
        except Exception as exc:
            title = f"(parse err: {exc})"
        anchor = getattr(ch, "anchor", None)
        pos = ""
        if anchor and hasattr(anchor, "_from"):
            fr = anchor._from
            pos = f"col {fr.col + 1} row {fr.row + 1}"
        lines.append(f"  {i}. Title: {title} | Position: {pos}")

    lines.append("")
    lines.append("=== CHART SERIES DATA (from formulas in 3 GRAPHS if any) ===")
    wb_f = load_workbook(PATH, data_only=False)
    wsf = wb_f["3 GRAPHS"]
    for r in range(1, min(wsf.max_row, 310) + 1):
        for c in range(1, min(wsf.max_column, 40) + 1):
            v = wsf.cell(r, c).value
            if isinstance(v, str) and v.startswith("=") and "MAGIC NUMBERS" in v.upper():
                lines.append(f"{get_column_letter(c)}{r}: {v[:120]}")

    # Scan for chart title text in merged/header cells row 5-15 per section
    lines.append("")
    lines.append("=== HEADER ROWS D-E (possible chart titles) ===")
    for r in range(1, min(ws.max_row, 310) + 1):
        parts = []
        for c in range(4, 12):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and len(v.strip()) > 2:
                parts.append(f"{get_column_letter(c)}{r}={v.strip()[:60]}")
        if parts:
            lines.append(" | ".join(parts))

    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT}")
    print("\n".join(lines[:80]))


if __name__ == "__main__":
    main()
