from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH = r"c:\Users\HOME\Desktop\4 - FUND JAIRO\2. ASSETS\2.1 THESIS\0. THESIS TOOL\FINANCIAL TOOL.xlsx"
wb = load_workbook(PATH, data_only=False)

# 1 DATA structure - section headers in col B
ws = wb["1 DATA"]
print("=== 1 DATA - ALL ROW LABELS (col B) ===")
for r in range(2, 250):
    v = ws.cell(r, 2).value
    if v:
        print(f"B{r}={v}")

# Year headers row 12-13?
print("\n=== 1 DATA year columns (row 12-13, cols E-P) ===")
for c in range(5, 17):
    print(get_column_letter(c), ws.cell(12, c).value, ws.cell(13, c).value)

# Bloomberg formulas in 0 get DATA - sample C10
ws0 = wb["0 get DATA"]
print("\n=== 0 get DATA - Bloomberg formula samples ===")
samples = ["C10", "E14", "E99", "AM15", "AA16", "AP14", "F11"]
for addr in samples:
    cell = ws0[addr]
    print(addr, cell.value)

# Check columns AA-AP headers (multiples section)
print("\n=== 1 DATA multiples section (row 8-20, cols X-AP) ===")
for r in range(8, 21):
    row = []
    for c in range(24, 43):
        v = ws.cell(r, c).value
        if v:
            row.append(f"{get_column_letter(c)}{r}={v}")
    if row:
        print(" | ".join(row))

# 5 EST & CRISIS key labels
print("\n=== 5 EST & CRISIS (col B-D labels) ===")
ws5 = wb["5 EST & CRISIS"]
for r in range(1, 74):
    labels = [ws5.cell(r, c).value for c in range(1, 5)]
    if any(labels):
        print(r, labels)

# DCF section in METHOD VALUATION rows 40-69
print("\n=== 4 METHOD VALUATION rows 40-69 ===")
ws4 = wb["4 METHOD VALUATION"]
for r in range(40, 70):
    row = []
    for c in range(2, 20):
        v = ws4.cell(r, c).value
        if v:
            row.append(f"{get_column_letter(c)}{r}={str(v)[:80]}")
    if row:
        print(" | ".join(row))
