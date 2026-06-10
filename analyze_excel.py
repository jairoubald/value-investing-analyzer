from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH = r"c:\Users\HOME\Desktop\4 - FUND JAIRO\2. ASSETS\2.1 THESIS\0. THESIS TOOL\FINANCIAL TOOL.xlsx"


def dump_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        row = []
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if v is None:
                continue
            s = str(v).replace("\n", " ")[:100]
            row.append(f"{get_column_letter(c)}{r}={s}")
        if row:
            print(" | ".join(row))


def count_formulas(ws, max_row=200, max_col=50):
    formulas = []
    for r in range(1, min(max_row, ws.max_row) + 1):
        for c in range(1, min(max_col, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                formulas.append((get_column_letter(c) + str(r), v[:150]))
    return formulas


wb = load_workbook(PATH, data_only=False)

print("SHEETS:", wb.sheetnames)
print()

for sheet in wb.sheetnames:
    ws = wb[sheet]
    formulas = count_formulas(ws, max_row=ws.max_row, max_col=ws.max_column)
    print(f"=== {sheet} ({ws.max_row}x{ws.max_column}, formulas={len(formulas)}) ===")

sections = {
    "INTRO": (1, 13, 1, 12),
    "0 get DATA headers": ("0 get DATA", 1, 10, 2, 47),
    "1 DATA headers": ("1 DATA", 1, 10, 2, 47),
    "2 MAGIC NUMBERS": ("2 MAGIC NUMBERS", 1, 133, 1, 29),
    "3 GRAPHS labels": ("3 GRAPHS", 1, 50, 1, 39),
    "4 METHOD VALUATION": ("4 METHOD VALUATION", 1, 69, 1, 35),
    "5 EST & CRISIS": ("5 EST & CRISIS", 1, 74, 1, 72),
}

for label, spec in sections.items():
    if label == "INTRO":
        ws = wb["INTRO"]
        r1, r2, c1, c2 = spec
    else:
        sheet, r1, r2, c1, c2 = spec
        ws = wb[sheet]
    print(f"\n######## {label} ########")
    dump_range(ws, r1, r2, c1, c2)

print("\n######## SAMPLE FORMULAS (2 MAGIC NUMBERS) ########")
for addr, f in count_formulas(wb["2 MAGIC NUMBERS"])[:25]:
    print(f"{addr}: {f}")

print("\n######## SAMPLE FORMULAS (4 METHOD VALUATION) ########")
for addr, f in count_formulas(wb["4 METHOD VALUATION"])[:25]:
    print(f"{addr}: {f}")

print("\n######## SAMPLE FORMULAS (5 EST & CRISIS) ########")
for addr, f in count_formulas(wb["5 EST & CRISIS"])[:25]:
    print(f"{addr}: {f}")

# Row labels in 1 DATA column B
print("\n######## 1 DATA - row labels (col B, rows 2-80) ########")
ws = wb["1 DATA"]
for r in range(2, 81):
    v = ws.cell(r, 2).value
    if v:
        print(f"B{r}={v}")

# Check input cells in 0 get DATA
print("\n######## 0 get DATA - input area (B2:F10) ########")
dump_range(wb["0 get DATA"], 2, 10, 2, 10)
